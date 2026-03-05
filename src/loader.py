"""
loader.py  (PRODUCTION REWRITE)
────────────────────────────────
Key fixes:
  ✅  Generator-based streaming (no full-file memory load)
  ✅  Deduplication by conversation ID
  ✅  Per-record validation via Pydantic (bad records logged, not silently dropped)
  ✅  Handles JSONL, JSON, CSV, XLSX
  ✅  Logs dropped count at every stage
"""
from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Generator, Iterator

from .models import Conversation, Message

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_conversation(raw: dict, source_file: str = "") -> Conversation | None:
    """Parse raw dict into a validated Conversation. Returns None on failure."""
    try:
        messages_raw = raw.get("messages", [])
        messages = []
        for i, m in enumerate(messages_raw):
            try:
                messages.append(Message(**m))
            except Exception as me:
                logger.warning(f"[{source_file}] Skipping malformed message {i}: {me}")

        if not messages:
            logger.warning(f"[{source_file}] Conversation '{raw.get('id')}' has no valid messages — skipped")
            return None

        return Conversation(
            id=str(raw.get("id", "")),
            messages=messages,
            intent=raw.get("intent"),
            issue=raw.get("issue"),
            issue_type=raw.get("issue_type"),
            is_escalated=bool(raw.get("is_escalated", False)),
            started_at=raw.get("started_at"),
            metadata={k: v for k, v in raw.items()
                      if k not in {"id", "messages", "intent", "issue", "issue_type",
                                   "is_escalated", "started_at"}},
        )
    except Exception as e:
        logger.error(f"[{source_file}] Failed to parse conversation: {e} | raw keys={list(raw.keys())}")
        return None


def _dedup_stream(
    stream: Iterator[Conversation],
) -> Generator[Conversation, None, None]:
    """Yield conversations, skipping duplicate IDs. Logs count at end."""
    seen: set[str] = set()
    total = dupes = 0
    for conv in stream:
        total += 1
        if conv.id in seen:
            dupes += 1
            continue
        seen.add(conv.id)
        yield conv
    if dupes:
        logger.info(f"Deduplication: {dupes}/{total} duplicates removed ({total - dupes} kept)")


# ── JSONL loader ──────────────────────────────────────────────────────────────

def load_conversations_from_jsonl(
    directory: Path,
) -> Generator[Conversation, None, None]:
    """Stream conversations line-by-line from all *.jsonl files in directory."""
    files = sorted(directory.glob("*.jsonl"))
    if not files:
        logger.warning(f"No .jsonl files found in {directory}")
        return

    def _stream():
        for path in files:
            logger.info(f"Loading JSONL: {path.name}")
            ok = skipped = 0
            with open(path, encoding="utf-8") as f:
                for lineno, line in enumerate(f, 1):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        raw = json.loads(line)
                    except json.JSONDecodeError as e:
                        logger.warning(f"{path.name}:{lineno} JSON parse error: {e}")
                        skipped += 1
                        continue
                    conv = _make_conversation(raw, source_file=path.name)
                    if conv:
                        ok += 1
                        yield conv
                    else:
                        skipped += 1
            logger.info(f"{path.name}: {ok} loaded, {skipped} skipped")

    yield from _dedup_stream(_stream())


# ── JSON loader ───────────────────────────────────────────────────────────────

def load_conversations_from_json(
    directory: Path,
) -> Generator[Conversation, None, None]:
    """Load conversations from JSON files (array at root or under a 'data' key)."""
    files = sorted(directory.glob("*.json"))
    if not files:
        logger.warning(f"No .json files found in {directory}")
        return

    def _stream():
        for path in files:
            logger.info(f"Loading JSON: {path.name}")
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception as e:
                logger.error(f"Failed to read {path.name}: {e}")
                continue

            records = data if isinstance(data, list) else data.get("data", [])
            ok = skipped = 0
            for raw in records:
                conv = _make_conversation(raw, source_file=path.name)
                if conv:
                    ok += 1
                    yield conv
                else:
                    skipped += 1
            logger.info(f"{path.name}: {ok} loaded, {skipped} skipped")

    yield from _dedup_stream(_stream())


# ── CSV loader ────────────────────────────────────────────────────────────────

def load_conversations_from_csv(
    directory: Path,
    column_mapping: dict[str, str],   # e.g. {"conversation_id": "id", "text": "messages"}
) -> Generator[Conversation, None, None]:
    """Load flat-format CSV. Each row = one conversation (messages as a single text field)."""
    files = sorted(directory.glob("*.csv"))
    if not files:
        logger.warning(f"No .csv files found in {directory}")
        return

    def _stream():
        for path in files:
            logger.info(f"Loading CSV: {path.name}")
            ok = skipped = 0
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for rownum, row in enumerate(reader, 1):
                    # Remap columns
                    mapped = {column_mapping.get(k, k): v for k, v in row.items()}
                    # Wrap flat text into message list
                    raw_text = mapped.pop("messages", mapped.pop("text", "")).strip()
                    if not raw_text:
                        logger.warning(f"{path.name}:{rownum} Empty text — skipped")
                        skipped += 1
                        continue
                    mapped["messages"] = [{"role": "user", "text": raw_text}]
                    if not mapped.get("id"):
                        mapped["id"] = f"{path.stem}_{rownum}"
                    conv = _make_conversation(mapped, source_file=path.name)
                    if conv:
                        ok += 1
                        yield conv
                    else:
                        skipped += 1
            logger.info(f"{path.name}: {ok} loaded, {skipped} skipped")

    yield from _dedup_stream(_stream())


# ── XLSX loader ───────────────────────────────────────────────────────────────

def load_conversations_from_xlsx(
    directory: Path,
    column_mapping: dict[str, str],
    sheet_name: str = "Data",
) -> Generator[Conversation, None, None]:
    """Load from Excel. Requires openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl --break-system-packages")

    files = sorted(directory.glob("*.xlsx"))
    if not files:
        logger.warning(f"No .xlsx files found in {directory}")
        return

    def _stream():
        for path in files:
            logger.info(f"Loading XLSX: {path.name} sheet='{sheet_name}'")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                logger.error(f"{path.name}: sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                continue
            ws = wb[sheet_name]
            rows = iter(ws.iter_rows(values_only=True))
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows, []))]
            ok = skipped = 0
            for rownum, row in enumerate(rows, 2):
                raw = {column_mapping.get(h, h): (v or "") for h, v in zip(headers, row)}
                raw_text = str(raw.pop("messages", raw.pop("text", ""))).strip()
                if not raw_text:
                    skipped += 1
                    continue
                raw["messages"] = [{"role": "user", "text": raw_text}]
                if not raw.get("id"):
                    raw["id"] = f"{path.stem}_{rownum}"
                conv = _make_conversation(raw, source_file=path.name)
                if conv:
                    ok += 1
                    yield conv
                else:
                    skipped += 1
            logger.info(f"{path.name}: {ok} loaded, {skipped} skipped")

    yield from _dedup_stream(_stream())
